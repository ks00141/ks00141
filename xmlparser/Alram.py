import os
import xml.etree.ElementTree as etree
import openpyxl
import sys
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.Qt import QFileSystemModel
from PyQt5.QtCore import *
from time import sleep

main_form,main_form_widget = uic.loadUiType("UI.ui")

class Thread(QThread):
    def __init__(self,parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        self.parent.created()

class XmlFile():
    def __init__(self,_path):
        self.xmlfile = etree.parse(_path)
        self.setTime()
        self.setDesc()

    def getTime(self):
        return self.time

    def setTime(self):
        self.time = self.xmlfile.find("/TimeStamp")

    def getDesc(self):
        return self.desc

    def setDesc(self):
        self.desc = self.xmlfile.find("/Description")

class Path():
    def __init__(self):
        self.SAVEPATH = "//10.21.10.204/fab 기술/fab기술/00_BackPart/MI/주차별 설비 알람"
        self.TEMPPATH = "//10.21.10.204/fab 기술/fab기술/00_BackPart/06_심영현/설비별 알람 리스트"
        self.ROOTPATH = "//10.21.10.204/fab 기술/fab기술/00_BackPart/06_심영현/설비별 알람 리스트/log_file"
        self.pathAddedWeek = None
        self.tools = None
        self.dateList = None

    def getRootPath(self):
        return self.ROOTPATH

    def setPathAddedDate(self,_date):
        self.date = _date

    def getPathAddedDate(self):
        return self.date

    def setPathAddedWeek(self,_PathAddedweek):
        self.pathAddedWeek = _PathAddedweek

    def getPathAddedWeek(self):
        return self.pathAddedWeek

    def setPathAddedTools(self,_tools):
        self.tools = _tools
        self.tools.sort(key=len)

class Xlsx():
    def __init__(self):   
        self.workBook = openpyxl.Workbook()
        self.workBook.create_sheet("#9",0)
        self.workBook.create_sheet("#10",1)
        self.workBook.create_sheet("#11",2)
        self.workBook.create_sheet("#12",3)

    def xlsxSet(self,_tool):
        self.worksheet = self.workBook[_tool]
        self.worksheet.column_dimensions['A'].width = 30
        self.worksheet.column_dimensions['B'].width = 150

    def xlsxInput(self,_time,_desc):
        self.worksheet.append([_time,_desc])
    
class Main(main_form,main_form_widget):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.show()
        self.xlsx = Xlsx()
        self.filepath = Path()
        self.xmlFile = None

        self.model = QFileSystemModel()
        self.model.setRootPath(self.filepath.ROOTPATH)
        self.listView.setModel(self.model)
        self.index_root = self.model.index(self.model.rootPath())
        self.listView.setRootIndex(self.index_root)
 

        self.createBtn.clicked.connect(self.threadAction)
        self.setBtn.clicked.connect(self.set)
        

    def  threadAction(self):
        x = Thread(self)
        x.start()

    def set(self):
        try:
            self.filepath.setPathAddedWeek(self.model.filePath(self.listView.currentIndex()))
            self.filepath.setPathAddedTools(os.listdir(self.filepath.getPathAddedWeek()))
            self.lblPath.setText(f'[ROOT PATH] / {self.listView.currentIndex().data()}')
        except:
            self.lblPath.setText("작업할 폴더를 선택해주세요.")
        

    def created(self):
        try:
            self.filepath.dateList = [os.listdir(f"{self.filepath.getPathAddedWeek()}/{tool}") for tool in self.filepath.tools]
            self.lblState.setText("작업중")
            sleep(1)
            for index in range(4):
                self.xlsx.xlsxSet(self.filepath.tools[index])
                for date in self.filepath.dateList[index]:
                    for file in os.listdir((f"{self.filepath.getPathAddedWeek()}/{self.filepath.tools[index]}/{date}")):
                        self.xmlFile = XmlFile(f"{self.filepath.getPathAddedWeek()}/{self.filepath.tools[index]}/{date}/{file}")
                        print(f"{self.xmlFile.getTime().text} _ {self.xmlFile.getDesc().text}")
                        self.xlsx.xlsxInput(self.xmlFile.getTime().text,self.xmlFile.getDesc().text)
            self.xlsx.workBook.save(f"{self.filepath.SAVEPATH}/{self.listView.currentIndex().data()}.xlsx")
            self.lblState.setText("완료")
        except:
            self.lblPath.setText("작업 폴더를 SET 해주세요.")


app = QApplication(sys.argv)
w = Main()
sys.exit(app.exec_())