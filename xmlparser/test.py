import openpyxl
import configparser

class ConfigRowString():
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('ROW_STRINg.ini',encoding='utf-8')

    def getFirstSection(self):
        try:
            returnValue = self.config.sections()[0]
        
        except:
            pass

        else:
            return returnValue

    def getValues(self,_section):
        values = list()
        try:
            for index in self.config[_section]:
                values.append(self.config[_section][index])

        except:
            pass
        
        else:
            return values

class Xlsx():
    def __init__(self):   
        self.workBook = openpyxl.Workbook()
        self.workBook.create_sheet("#9",0)
        self.workBook.create_sheet("#10",1)
        self.workBook.create_sheet("#11",2)
        self.workBook.create_sheet("#12",3)
        self.workBook.create_sheet("sum",4)
        self.configRowString = ConfigRowString()
        self.sumInit()
        self.workBook.save("./test.xlsx")

    def xlsxSet(self,_tool):
        self.worksheet = self.workBook[_tool]
        self.worksheet.column_dimensions['A'].width = 30
        self.worksheet.column_dimensions['B'].width = 150

    def xlsxInput(self,_time,_desc):
        self.worksheet.append([_time,_desc])

    def sumInit(self):
        self.worksheet = self.workBook["sum"]
        self.worksheet["g3"] = "sum"
        self.worksheet["b3"] = "Desciption"
        self.worksheet["c3"] = "tool"
        self.worksheet["c4"] = "#9"
        self.worksheet["d4"] = "#10"
        self.worksheet["e4"] = "#11"
        self.worksheet["f4"] = "#12"
        self.worksheet.merge_cells("b3:b4")
        self.worksheet.merge_cells("c3:f3")
        self.worksheet.merge_cells("g3:g4")
        self.worksheet["b3"].alignment = openpyxl.styles.Alignment(horizontal = "center",vertical = "center")
        self.worksheet["c3"].alignment = openpyxl.styles.Alignment(horizontal = "center",vertical = "center")
        self.worksheet["g3"].alignment = openpyxl.styles.Alignment(horizontal = "center",vertical = "center")
        self.worksheet.column_dimensions['B'].width = 100
        self.asterisk = '"*"'
        self.values = self.configRowString.getValues(self.configRowString.getFirstSection())
        for i,string in enumerate(self.values):
            self.worksheet[f"b{i+5}"] = string
            self.worksheet[f"c{i+5}"] = f"=COUNTIF('#9'!$B$2:$B$3000,$B{i+5} & {self.asterisk})"
            self.worksheet[f"d{i+5}"] = f"=COUNTIF('#10'!$B$2:$B$3000,$B{i+5} & {self.asterisk})"
            self.worksheet[f"e{i+5}"] = f"=COUNTIF('#11'!$B$2:$B$3000,$B{i+5} & {self.asterisk})"
            self.worksheet[f"f{i+5}"] = f"=COUNTIF('#12'!$B$2:$B$3000,$B{i+5} & {self.asterisk})"
            self.worksheet[f"g{i+5}"] = f"=SUM(C{i+5}:F{i+5})"
        
            


test = Xlsx()