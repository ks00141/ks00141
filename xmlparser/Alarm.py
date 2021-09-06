import os
import xml.etree.ElementTree as etree
import openpyxl
import sys
import configparser
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.Qt import QFileSystemModel
from PyQt5.QtCore import *


main_form,main_form_widget = uic.loadUiType("UI.ui")
stringset_form,stringset_form_widget = uic.loadUiType("stringset.ui")




class Thread(QThread):
    def __init__(self,parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        self.parent.created()

class XmlFile():
    def __init__(self,_path):
        self.xmlfile = etree.parse(_path)
        self.setTime()
        self.setDesc()

    def getTime(self):
        return self.time

    def setTime(self):
        self.time = self.xmlfile.find("./TimeStamp")

    def getDesc(self):
        return self.desc

    def setDesc(self):
        self.desc = self.xmlfile.find("./Description")

class Path():
    def __init__(self):
        self.SAVEPATH = "//10.21.10.204/fab 기술/fab기술/00_BackPart/MI/주차별 설비 알람"
        self.TEMPPATH = "//10.21.10.204/fab 기술/fab기술/00_BackPart/06_심영현/설비별 알람 리스트"
        self.ROOTPATH = "//10.21.10.204/fab 기술/fab기술/00_BackPart/06_심영현/설비별 알람 리스트/log_file"
        self.pathAddedWeek = None
        self.tools = None
        self.dateList = None

    def getRootPath(self):
        return self.ROOTPATH

    def setPathAddedDate(self,_date):
        self.date = _date

    def getPathAddedDate(self):
        return self.date

    def setPathAddedWeek(self,_PathAddedweek):
        self.pathAddedWeek = _PathAddedweek

    def getPathAddedWeek(self):
        return self.pathAddedWeek

    def setPathAddedTools(self,_tools):
        self.tools = _tools
        self.tools.sort(key=len)

class ConfigRowString():
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('ROW_STRINg.ini',encoding='utf-8')

    def getFirstSection(self):
        try:
            returnValue = self.config.sections()[0]
        
        except:
            pass

        else:
            return returnValue

    def getValues(self,_section):
        values = list()
        try:
            for index in self.config[_section]:
                values.append(self.config[_section][index])

        except:
            pass
        
        else:
            return values

    def getValuesIndex(self,_section):
        values = list()
        try:
            for index in self.config[_section]:
                values.append([index,self.config[_section][index]])

        except:
            pass
        
        else:
            return values

class StringSet(stringset_form,stringset_form_widget):
    def __init__(self):
        super().__init__()
        self.configrowString = ConfigRowString()
        self.setupUi(self)
        self.show()
        self.initString()
        # self.le.returnPressed.connect(self.append)
    
    def initString(self):
        self.values = self.configrowString.getValuesIndex(self.configrowString.getFirstSection())
        for value in self.values:
            self.tb.append(f"{value[0]}) {value[1]}")
    
    # def append(self):
    #     print(self.le.text())
    #     self.le.clear()


class Xlsx():
    def __init__(self):   
        self.workBook = openpyxl.Workbook()
        self.workBook.create_sheet("#9",0)
        self.workBook.create_sheet("#10",1)
        self.workBook.create_sheet("#11",2)
        self.workBook.create_sheet("#12",3)
        self.workBook.create_sheet("sum",4)
        self.configRowString = ConfigRowString()
        self.sumInit()
        self.workBook.save("./test.xlsx")

    def xlsxSet(self,_tool):
        self.worksheet = self.workBook[_tool]
        self.worksheet.column_dimensions['A'].width = 30
        self.worksheet.column_dimensions['B'].width = 150

    def xlsxInput(self,_time,_desc):
        self.worksheet.append([_time,_desc])

    def sumInit(self):
        self.worksheet = self.workBook["sum"]
        self.worksheet["g3"] = "sum"
        self.worksheet["b3"] = "Desciption"
        self.worksheet["c3"] = "tool"
        self.worksheet["c4"] = "#9"
        self.worksheet["d4"] = "#10"
        self.worksheet["e4"] = "#11"
        self.worksheet["f4"] = "#12"
        self.worksheet.merge_cells("b3:b4")
        self.worksheet.merge_cells("c3:f3")
        self.worksheet.merge_cells("g3:g4")
        self.worksheet["b3"].alignment = openpyxl.styles.Alignment(horizontal = "center",vertical = "center")
        self.worksheet["c3"].alignment = openpyxl.styles.Alignment(horizontal = "center",vertical = "center")
        self.worksheet["g3"].alignment = openpyxl.styles.Alignment(horizontal = "center",vertical = "center")
        self.worksheet["c3"].alignment = openpyxl.styles.Alignment(horizontal = "center",vertical = "center")
        self.worksheet["d3"].alignment = openpyxl.styles.Alignment(horizontal = "center",vertical = "center")
        self.worksheet["e3"].alignment = openpyxl.styles.Alignment(horizontal = "center",vertical = "center")
        self.worksheet["f3"].alignment = openpyxl.styles.Alignment(horizontal = "center",vertical = "center")
        self.worksheet.column_dimensions['B'].width = 100
        self.asterisk = '"*"'
        self.values = self.configRowString.getValues(self.configRowString.getFirstSection())
        for i,string in enumerate(self.values):
            self.worksheet[f"b{i+5}"] = string
            self.worksheet[f"c{i+5}"] = f"=COUNTIF('#9'!$B$2:$B$3000,$B{i+5} & {self.asterisk})"
            self.worksheet[f"d{i+5}"] = f"=COUNTIF('#10'!$B$2:$B$3000,$B{i+5} & {self.asterisk})"
            self.worksheet[f"e{i+5}"] = f"=COUNTIF('#11'!$B$2:$B$3000,$B{i+5} & {self.asterisk})"
            self.worksheet[f"f{i+5}"] = f"=COUNTIF('#12'!$B$2:$B$3000,$B{i+5} & {self.asterisk})"
            self.worksheet[f"g{i+5}"] = f"=SUM(C{i+5}:F{i+5})"
          
class Main(main_form,main_form_widget):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.show()
        self.filepath = Path()
        self.xmlFile = None
        self.workState = [".","..","...","....",".....","......"," .....","  ....","   ...","    ..","     ."]

        self.model = QFileSystemModel()
        self.model.setRootPath(self.filepath.ROOTPATH)
        self.listView.setModel(self.model)
        self.index_root = self.model.index(self.model.rootPath())
        self.listView.setRootIndex(self.index_root)
 

        self.createBtn.clicked.connect(self.threadAction)
        self.setBtn.clicked.connect(self.set)
        self.stringset.triggered.connect(self.stringsetAction)
        

    def  threadAction(self):
        x = Thread(self)
        x.start()

    def set(self):
        
        try:
            self.filepath.setPathAddedWeek(self.model.filePath(self.listView.currentIndex()))
            self.filepath.setPathAddedTools(os.listdir(self.filepath.getPathAddedWeek()))
            self.lblPath.setText(f'[ROOT PATH] / {self.listView.currentIndex().data()}')
        except:
            self.lblPath.setText("작업할 폴더를 선택해주세요.")
        

    def created(self):
        self.xlsx = Xlsx()
        try:
            self.filepath.dateList = [os.listdir(f"{self.filepath.getPathAddedWeek()}/{tool}") for tool in self.filepath.tools]
            for index in range(4):
                self.xlsx.xlsxSet(self.filepath.tools[index])
                for date in self.filepath.dateList[index]:
                    for fileIndex,file in enumerate(os.listdir((f"{self.filepath.getPathAddedWeek()}/{self.filepath.tools[index]}/{date}"))):
                        self.lblState.setText(self.workState[fileIndex%7])
                        self.xmlFile = XmlFile(f"{self.filepath.getPathAddedWeek()}/{self.filepath.tools[index]}/{date}/{file}")
                        self.xlsx.xlsxInput(self.xmlFile.getTime().text,self.xmlFile.getDesc().text)
            self.xlsx.workBook.save(f"{self.filepath.SAVEPATH}/{self.listView.currentIndex().data()}.xlsx")
            self.lblState.setText("완료")
        except:
            self.lblPath.setText("작업 폴더를 SET 해주세요.")

    def stringsetAction(self):
        strigset = StringSet()
        strigset.exec_()

app = QApplication(sys.argv)
w = Main()
sys.exit(app.exec_())