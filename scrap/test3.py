# selenium webdriver
# keys
# datetime
# time
# openpyxl

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import datetime
import time
import openpyxl

# 현재시간 (파일명 저장용도, yymmdd)
now = datetime.datetime.now()
date = now.strftime('%y%m%d')

# NS = CSP,WLP / TS = TC-CSP, TC-WLP / PST = PST AOI
TC = '[최종외검][TC-CSP,TC-WLP]최종외검 저수율 리스트'
NS = '[최종외검][CSP,WLP]최종외검 저수율 리스트'
PST = '[PST Map]PST AOI 저수율 리스트'

# wisol 그룹웨어 url
url = 'http://gw.wisol.co.kr'

# Chrome Webdriver 경로
url_driver = 'C:/Users/wisol/Desktop/chromedriver.exe'

# Webdriver 가동
driver = webdriver.Chrome(url_driver)
driver.get(url)

# 그룹웨어 Login page ID / PWD 입력창 가져오기
id = driver.find_element_by_id('id')
pwd = driver.find_element_by_id('password')

# ID / PWD 입력값 보내기 / Keys.RETURN : ENTER Key 입력
id.send_keys('yhsim')
pwd.send_keys('wisol12')
pwd.send_keys(Keys.RETURN)

# 그룹웨어 Main Page 진입, POPUP Window 기다리기 (1초)
time.sleep(1)

# POPUP Window 전환후 종료 -> Main Window 전환
driver.switch_to_window(driver.window_handles[1])
driver.close()
driver.switch_to_window(driver.window_handles[0])

# Frame 구조

# 1 Depth
#  duplChkFrame : 404일때 나오는 페이지 인듯
#  protalBody : Main Body
#  #wrapMainFix -> .gnbContainer -> .lnbContainer : 상단 Manu
#   lnbContainer List
#    li:nth-child(1) : 전자우편
#    li:nth-child(2) : 게시판
#    li:nth-child(3) : 전자결재
#    li:nth-child(4) : PIMS
#    li:nth-child(5) : 문서함
#    li:nth-child(6) : 자원예약
#    li:nth-child(7) : 파일함
#    li:nth-child(8) : Wisol Square
#    li:nth-child(9) : 세미나

# 2 Depth
#  idHeadTopIfm : ?
#  idHeadTopFixedCountIfm : ?
#  subBody : 하단 Side Menu + Contents body

# 3 Depth
#  menuFrame : Side Menu
#  contentFrame : contents

driver.switch_to_frame('portalBody')
driver.find_element_by_css_selector('#wrapMainFix > div.lnbContainer > div > div.nav > ul > li:nth-child(1) > a').click()
driver.switch_to_frame('subBody')
driver.switch_to_frame('menuFrame')
driver.find_element_by_css_selector('#ext-gen10 > div > li:nth-child(4) > ul > li > div > a').click()


date = input('조회일자 : ')

if date == '':
    date = datetime.datetime.now().strftime("[%Y-%m-%d]")
else:
    date = f'[20{date[0:2]}-{date[2:4]}-{date[4:6]}]'

group = [f'{TC} {date}',f'{NS} {date}',f'{PST} {date}']

for i in range(3):

    elem = None
    driver.switch_to.default_content()
    driver.switch_to_frame('portalBody')
    driver.switch_to_frame('subBody')
    driver.switch_to_frame('contentFrame')
    mail_subject = driver.find_elements_by_class_name('subject div a')

    for subject in mail_subject:

        if str(subject.get_attribute('innerHTML')).find('strong') > 0:
            elem = subject.find_element_by_xpath('./strong')
        else:
            elem = subject

        if elem.get_attribute('innerHTML') == group[i]:
            elem.click()
            summary = driver.find_element_by_id('summary')
            tr = summary.find_elements_by_xpath('./tbody/tr')[1:]
            if group[i] == f'{TC} {date}':
                print(group[i])
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1).value = 'NO'
                sheet.cell(row=1, column=2).value = '기종'
                sheet.cell(row=1, column=3).value = 'LOT NO'
                sheet.cell(row=1, column=4).value = 'WAFER ID(MAP)'
                sheet.cell(row=1, column=5).value = '수율(%)'
                for i in tr:
                    td = i.find_elements_by_xpath('./td')
                    no = int(td[0].get_attribute('innerHTML'))
                    lotid = td[1].get_attribute('innerHTML')
                    waferid_link = td[2].find_element_by_xpath('./a').get_attribute('href')
                    waferid = td[2].find_element_by_xpath('./a').get_attribute('innerHTML')
                    device = td[4].get_attribute('innerHTML')
                    yield_ = float(td[6].get_attribute('innerHTML'))
                    sheet.cell(row=no+1,column=1).value = no
                    sheet.cell(row=no+1,column=2).value = device
                    sheet.cell(row=no+1,column=3).value = lotid
                    sheet.cell(row=no+1,column=4).value = waferid
                    sheet.cell(row=no+1,column=4).hyperlink = waferid_link
                    sheet.cell(row=no+1,column=4).style = 'Hyperlink'
                    sheet.cell(row=no+1,column=5).value = yield_
                wb.save(f'./저수율list/TC/TC_저수율list_{date}.xlsx')
                print('작업 완료')
                list_btn = driver.find_element_by_css_selector('#wrap > form > div:nth-child(20) > ul.menuLeft > li:nth-child(8) > a > span')
                list_btn.click()
                break
            elif group[i] == f'{NS} {date}':
                print(group[i])
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1).value = 'NO'
                sheet.cell(row=1, column=2).value = '기종'
                sheet.cell(row=1, column=3).value = 'LOT NO'
                sheet.cell(row=1, column=4).value = 'WAFER ID(MAP)'
                sheet.cell(row=1, column=5).value = '수율(%)'
                for i in tr:
                    td = i.find_elements_by_xpath('./td')
                    no = int(td[0].get_attribute('innerHTML'))
                    lotid = td[1].get_attribute('innerHTML')
                    waferid_link = td[2].find_element_by_xpath('./a').get_attribute('href')
                    waferid = td[2].find_element_by_xpath('./a').get_attribute('innerHTML')
                    device = td[4].get_attribute('innerHTML')
                    yield_ = float(td[7].get_attribute('innerHTML'))
                    sheet.cell(row=no+1,column=1).value = no
                    sheet.cell(row=no+1,column=2).value = device
                    sheet.cell(row=no+1,column=3).value = lotid
                    sheet.cell(row=no+1,column=4).value = waferid
                    sheet.cell(row=no+1,column=4).hyperlink = waferid_link
                    sheet.cell(row=no+1,column=4).style = 'Hyperlink'
                    sheet.cell(row=no+1,column=5).value = yield_
                wb.save(f'./저수율list/NS/NS_저수율list_{date}.xlsx')
                print('작업 완료')
                list_btn = driver.find_element_by_css_selector('#wrap > form > div:nth-child(20) > ul.menuLeft > li:nth-child(8) > a > span')
                list_btn.click()
                break
            elif group[i] == f'{PST} {date}':
                print(group[i])
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1).value = 'NO'
                sheet.cell(row=1, column=2).value = '기종'
                sheet.cell(row=1, column=3).value = 'LOT NO'
                sheet.cell(row=1, column=4).value = 'WAFER ID(MAP)'
                sheet.cell(row=1, column=5).value = '수율(%)'
                for i in tr:
                    td = i.find_elements_by_xpath('./td')
                    no = int(td[0].get_attribute('innerHTML'))
                    lotid = td[1].get_attribute('innerHTML')
                    waferid_link = td[2].find_element_by_xpath('./a').get_attribute('href')
                    waferid = td[2].find_element_by_xpath('./a').get_attribute('innerHTML')
                    device = td[5].get_attribute('innerHTML')
                    yield_ = float(td[10].get_attribute('innerHTML'))
                    sheet.cell(row=no+1,column=1).value = no
                    sheet.cell(row=no+1,column=2).value = device
                    sheet.cell(row=no+1,column=3).value = lotid
                    sheet.cell(row=no+1,column=4).value = waferid
                    sheet.cell(row=no+1,column=4).hyperlink = waferid_link
                    sheet.cell(row=no+1,column=4).style = 'Hyperlink'
                    sheet.cell(row=no+1,column=5).value = yield_
                wb.save(f'./저수율list/PST/PST_저수율list_{date}.xlsx')
                print('작업 완료')
                list_btn = driver.find_element_by_css_selector('#wrap > form > div:nth-child(20) > ul.menuLeft > li:nth-child(8) > a > span')
                list_btn.click()
                break
        else:
            print('조회 실패')
