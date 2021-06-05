
import sys
import time
import datetime
import openpyxl
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QRadioButton, QButtonGroup, QStatusBar
from PyQt5.QtCore import QCoreApplication
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        
        self.setWindowTitle("test")
        self.setGeometry(1000,450,300,130)

        self.statusBar().showMessage("Ready")


        self.rbtn1 = QRadioButton("All",self)
        self.rbtn1.move(20,20)
        self.rbtn1.clicked.connect(self.each_clickcheck)
        self.rbtn1.clicked.connect(lambda  : self.statusBar().showMessage("All Select"))
        self.rbtn1.setChecked(True)

        self.rbtn2 = QRadioButton("TC",self)
        self.rbtn2.move(20,40)
        self.rbtn2.setAutoExclusive(False)
        self.rbtn2.clicked.connect(self.all_clickcheck)
        self.rbtn2.clicked.connect(lambda  : self.statusBar().showMessage("TC Select"))

        self.rbtn3 = QRadioButton("NS",self)
        self.rbtn3.move(20,60)
        self.rbtn3.setAutoExclusive(False)
        self.rbtn3.clicked.connect(self.all_clickcheck)
        self.rbtn3.clicked.connect(lambda  : self.statusBar().showMessage("NS Select"))

        self.rbtn4 = QRadioButton("PST",self)
        self.rbtn4.move(20,80)
        self.rbtn4.setAutoExclusive(False)
        self.rbtn4.clicked.connect(self.all_clickcheck)
        self.rbtn4.clicked.connect(lambda  : self.statusBar().showMessage("PST Select"))

        self.btn2 = QPushButton("open GW",self)
        self.btn2.resize(100,40)
        self.btn2.move(100,45)
        self.btn2.clicked.connect(self.openWeb)

        self.show()
    
    def all_clickcheck(self):
        if self.rbtn1.isChecked():
            self.rbtn1.setChecked(False)

    def each_clickcheck(self):
        self.each_check(self.rbtn2)
        self.each_check(self.rbtn3)
        self.each_check(self.rbtn4)
    
    def each_check(self,btn):
        if btn.isChecked():
            btn.setChecked(False)

    def scrap(self,driver):
        mode = None
        TC = '[최종외검][TC-CSP,TC-WLP]최종외검 저수율 리스트'
        NS = '[최종외검][CSP,WLP]최종외검 저수율 리스트'
        PST = '[PST Map]PST AOI 저수율 리스트'
        date = datetime.datetime.now().strftime("[%Y-%m-%d]")
        group = [f'{TC} {date}',f'{NS} {date}',f'{PST} {date}']
        
        if self.rbtn1.isChecked():
            self.TC_find(driver,group[0])
            self.NS_find(driver,group[1])
            self.PST_find(driver,group[2])
            print("작업 완료")
        else:
            if self.rbtn2.isChecked():
                self.TC_find(driver,group[0])
            if self.rbtn3.isChecked():
                self.NS_find(driver,group[1])
            if self.rbtn4.isChecked():
                self.PST_find(driver,group[2])
            print("작업 완료")
        driver.quit()

    def openWeb(self):
        webdriver_options = webdriver.ChromeOptions()
        webdriver_options .add_argument('headless')
        chromedriver = './chromedriver.exe'
        driver = webdriver.Chrome(chromedriver, options=webdriver_options )
        driver.get('http://gw.wisol.co.kr')
        id = driver.find_element_by_id('id')
        pwd = driver.find_element_by_id('password')
        id.send_keys("yhsim")
        pwd.send_keys("wisol1")
        pwd.send_keys(Keys.RETURN)
        
        self.popup_close(driver)
        self.goto_mail(driver)
        self.scrap(driver)

    def popup_close(self,driver):
        time.sleep(2)
        driver.switch_to_window(driver.window_handles[1])
        driver.close()
        driver.switch_to_window(driver.window_handles[0])

    def goto_mail(self,driver):
        driver.switch_to_frame('portalBody')
        driver.find_element_by_css_selector('#wrapMainFix > div.lnbContainer > div > div.nav > ul > li:nth-child(1) > a').click()
        driver.switch_to_frame('subBody')
        driver.switch_to_frame('menuFrame')
        driver.find_element_by_css_selector('#ext-gen10 > div > li:nth-child(4) > ul > li > div > a').click()
        driver.switch_to.default_content()
        driver.switch_to_frame('portalBody')
        driver.switch_to_frame('subBody')
        driver.switch_to_frame('contentFrame')

    def TC_find(self,driver,group):
        self.statusBar().showMessage("TC Scrapping")
        self.statusBar().repaint()
        elem = None
        driver.switch_to.default_content()
        driver.switch_to_frame('portalBody')
        driver.switch_to_frame('subBody')
        driver.switch_to_frame('contentFrame')
        mail_subject = driver.find_elements_by_class_name('subject div a')

        for subject in mail_subject:
            if str(subject.get_attribute('innerHTML')).find('strong') > 0:
                elem = subject.find_element_by_xpath('./strong')
            else:
                elem = subject

            if elem.get_attribute('innerHTML') == group:
                elem.click()
                summary = driver.find_element_by_id('summary')
                tr = summary.find_elements_by_xpath('./tbody/tr')[1:]
                print(group)
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1).value = 'NO'
                sheet.cell(row=1, column=2).value = '기종'
                sheet.cell(row=1, column=3).value = 'LOT NO'
                sheet.cell(row=1, column=4).value = 'WAFER ID(MAP)'
                sheet.cell(row=1, column=5).value = '수율(%)'
                for i in tr:
                    td = i.find_elements_by_xpath('./td')
                    no = int(td[0].get_attribute('innerHTML'))
                    lotid = td[1].get_attribute('innerHTML')
                    waferid_link = td[2].find_element_by_xpath('./a').get_attribute('href')
                    waferid = td[2].find_element_by_xpath('./a').get_attribute('innerHTML')
                    device = td[4].get_attribute('innerHTML')
                    yield_ = float(td[6].get_attribute('innerHTML'))
                    sheet.cell(row=no+1,column=1).value = no
                    sheet.cell(row=no+1,column=2).value = device
                    sheet.cell(row=no+1,column=3).value = lotid
                    sheet.cell(row=no+1,column=4).value = waferid
                    sheet.cell(row=no+1,column=4).hyperlink = waferid_link
                    sheet.cell(row=no+1,column=4).style = 'Hyperlink'
                    sheet.cell(row=no+1,column=5).value = yield_
                    self.statusBar().showMessage(f'TC {no} {device}_{lotid}_{waferid}_{yield_}')
                    self.statusBar().repaint()
                wb.save(f'C:/Users/wisol/Desktop/opencv/저수율list/TC/TC_저수율list_{group[-12:]}.xlsx')
                self.statusBar().showMessage('END')
                self.statusBar().repaint()
                list_btn = driver.find_element_by_css_selector('#wrap > form > div:nth-child(20) > ul.menuLeft > li:nth-child(8) > a > span')
                list_btn.click()
                break
                
    def NS_find(self,driver,group):
        self.statusBar().showMessage("NS Scrapping")
        self.statusBar().repaint()
        elem = None
        driver.switch_to.default_content()
        driver.switch_to_frame('portalBody')
        driver.switch_to_frame('subBody')
        driver.switch_to_frame('contentFrame')
        mail_subject = driver.find_elements_by_class_name('subject div a')

        for subject in mail_subject:
            if str(subject.get_attribute('innerHTML')).find('strong') > 0:
                elem = subject.find_element_by_xpath('./strong')
            else:
                elem = subject

            if elem.get_attribute('innerHTML') == group:
                elem.click()
                summary = driver.find_element_by_id('summary')
                tr = summary.find_elements_by_xpath('./tbody/tr')[1:]
                print(group)
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1).value = 'NO'
                sheet.cell(row=1, column=2).value = '기종'
                sheet.cell(row=1, column=3).value = 'LOT NO'
                sheet.cell(row=1, column=4).value = 'WAFER ID(MAP)'
                sheet.cell(row=1, column=5).value = '수율(%)'
                for i in tr:
                    td = i.find_elements_by_xpath('./td')
                    no = int(td[0].get_attribute('innerHTML'))
                    lotid = td[1].get_attribute('innerHTML')
                    waferid_link = td[2].find_element_by_xpath('./a').get_attribute('href')
                    waferid = td[2].find_element_by_xpath('./a').get_attribute('innerHTML')
                    device = td[4].get_attribute('innerHTML')
                    yield_ = float(td[7].get_attribute('innerHTML'))
                    sheet.cell(row=no+1,column=1).value = no
                    sheet.cell(row=no+1,column=2).value = device
                    sheet.cell(row=no+1,column=3).value = lotid
                    sheet.cell(row=no+1,column=4).value = waferid
                    sheet.cell(row=no+1,column=4).hyperlink = waferid_link
                    sheet.cell(row=no+1,column=4).style = 'Hyperlink'
                    sheet.cell(row=no+1,column=5).value = yield_
                    self.statusBar().showMessage(f'NS {no} {device}_{lotid}_{waferid}_{yield_}')
                    self.statusBar().repaint()
                wb.save(f'C:/Users/wisol/Desktop/opencv/저수율list/NS/NS_저수율list_{group[-12:]}.xlsx')
                self.statusBar().showMessage('END')
                self.statusBar().repaint()
                list_btn = driver.find_element_by_css_selector('#wrap > form > div:nth-child(20) > ul.menuLeft > li:nth-child(8) > a > span')
                list_btn.click()
                break

    def PST_find(self,driver,group):
        self.statusBar().showMessage("PST Scrapping")
        self.statusBar().repaint()
        elem = None
        driver.switch_to.default_content()
        driver.switch_to_frame('portalBody')
        driver.switch_to_frame('subBody')
        driver.switch_to_frame('contentFrame')
        mail_subject = driver.find_elements_by_class_name('subject div a')

        for subject in mail_subject:
            if str(subject.get_attribute('innerHTML')).find('strong') > 0:
                elem = subject.find_element_by_xpath('./strong')
            else:
                elem = subject

            if elem.get_attribute('innerHTML') == group:
                elem.click()
                summary = driver.find_element_by_id('summary')
                tr = summary.find_elements_by_xpath('./tbody/tr')[1:]
                print(group)
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1).value = 'NO'
                sheet.cell(row=1, column=2).value = '기종'
                sheet.cell(row=1, column=3).value = 'LOT NO'
                sheet.cell(row=1, column=4).value = 'WAFER ID(MAP)'
                sheet.cell(row=1, column=5).value = '수율(%)'
                for i in tr:
                    td = i.find_elements_by_xpath('./td')
                    no = int(td[0].get_attribute('innerHTML'))
                    lotid = td[1].get_attribute('innerHTML')
                    waferid_link = td[2].find_element_by_xpath('./a').get_attribute('href')
                    waferid = td[2].find_element_by_xpath('./a').get_attribute('innerHTML')
                    device = td[5].get_attribute('innerHTML')
                    yield_ = float(td[10].get_attribute('innerHTML'))
                    sheet.cell(row=no+1,column=1).value = no
                    sheet.cell(row=no+1,column=2).value = device
                    sheet.cell(row=no+1,column=3).value = lotid
                    sheet.cell(row=no+1,column=4).value = waferid
                    sheet.cell(row=no+1,column=4).hyperlink = waferid_link
                    sheet.cell(row=no+1,column=4).style = 'Hyperlink'
                    sheet.cell(row=no+1,column=5).value = yield_
                    self.statusBar().showMessage(f'PST {no} {device}_{lotid}_{waferid}_{yield_}')
                    self.statusBar().repaint()
                wb.save(f'C:/Users/wisol/Desktop/opencv/저수율list/PST/PST_저수율list_{group[-12:]}.xlsx')
                self.statusBar().showMessage('END')
                self.statusBar().repaint()
                list_btn = driver.find_element_by_css_selector('#wrap > form > div:nth-child(20) > ul.menuLeft > li:nth-child(8) > a > span')
                list_btn.click()
                break

if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = App()
    sys.exit(app.exec_())